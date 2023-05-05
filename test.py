from itertools import count
from openpyxl import Workbook, load_workbook


class ExportExcel:

    def __init__(self, filename=None, write_only=False):
        try:
            if filename:
                self.file_name = filename
                self.wb = load_workbook(filename)
            else:
                self.wb = Workbook(write_only=write_only)
        except FileNotFoundError:
            self.wb = Workbook(write_only=write_only)
        self.write_only = write_only
    
    def write_to_sheet(self, sheetname, title, data):
        ws = self.wb.create_sheet(title=sheetname)
        ws.append(title)
        for item in data:
            ws.append(item)
    
    def append_data_to_sheet(self, sheetname, data):
        ws = self.wb[sheetname]
        ws.append(data)
    
    def save(self, filename):
        self.wb.save(filename)
        self.wb.close()

    def get_sheet_data(self, sheet_name, is_title=True):
        data = []
        sheet = self.wb.get_sheet_by_name(sheet_name)

        for index, row in zip(count(), sheet.rows):
            if is_title and index == 0:
                continue
            temp_data = []
            for cell in row:
                temp_data.append(cell.value)
            
            data.append(temp_data)

        return data
